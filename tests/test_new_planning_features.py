import csv
import json
from pathlib import Path

import numpy as np
from openpyxl import Workbook

from app.services.assembly_path import (
    CapsuleCollisionWorld,
    PoseState,
    _segment_distances_batch,
    _transform_axis,
    plan_assembly_paths,
)
from app.services.ifc_geometry import Rebar
from app.services.planner import plan_installation, save_planning_outputs
from app.services.robot_path import generate_robot_outputs
from app.services.sequence_io import load_manual_sequence


def make_bar(index: int, points, radius: float = 5.0) -> Rebar:
    axis = np.asarray(points, dtype=float)
    return Rebar(
        index=index,
        entity_id=100 + index,
        guid=f"guid-{index}",
        name=f"bar-{index}",
        tag=f"B{index}",
        map_id=0,
        axis=axis,
        radius=radius,
        bbox_min=axis.min(0) - radius,
        bbox_max=axis.max(0) + radius,
        length=float(np.linalg.norm(np.diff(axis, axis=0), axis=1).sum()),
    )


def test_segment_distance_3d_crossing():
    distance = _segment_distances_batch(
        np.array([[0.0, 0.0, 0.0]]),
        np.array([[10.0, 0.0, 0.0]]),
        np.array([[5.0, -2.0, 0.0]]),
        np.array([[5.0, 2.0, 0.0]]),
    )
    assert distance[0] < 1e-9


def test_rigid_pose_rotates_rebar_about_pivot():
    bar = make_bar(0, [[-1, 0, 0], [1, 0, 0]])
    pose = PoseState(
        np.array([10.0, 0.0, 0.0]),
        np.array([0.0, 0.0, np.sin(np.pi / 4), np.cos(np.pi / 4)]),
    )
    transformed = _transform_axis(bar, np.zeros(3), pose)
    np.testing.assert_allclose(transformed, [[10, -1, 0], [10, 1, 0]], atol=1e-7)


def test_manual_xlsx_sequence_resolves_and_reorders(tmp_path: Path):
    bars = [
        make_bar(0, [[0, 0, 0], [100, 0, 0]]),
        make_bar(1, [[0, 100, 0], [100, 100, 0]]),
    ]
    path = tmp_path / "sequence.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["installation_step", "guid", "direction_x", "direction_y", "direction_z"])
    sheet.append([2, "guid-0", 1, 0, 0])
    sheet.append([1, "guid-1", None, None, None])
    workbook.save(path)

    sequence, stats = load_manual_sequence(path, bars)
    assert [row["bar_index"] for row in sequence] == [1, 0]
    assert sequence[1]["entry_direction"] == [1.0, 0.0, 0.0]
    assert stats["sequence_source"] == "excel"




def test_manual_xlsx_id_only_uses_row_order(tmp_path: Path):
    bars = [
        make_bar(0, [[0, 0, 0], [100, 0, 0]]),
        make_bar(1, [[0, 100, 0], [100, 100, 0]]),
        make_bar(2, [[0, 200, 0], [100, 200, 0]]),
    ]
    path = tmp_path / "id_sequence.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["\u94a2\u7b4bID"])
    sheet.append(["guid-2"])
    sheet.append(["guid-0"])
    sheet.append(["guid-1"])
    workbook.save(path)

    sequence, stats = load_manual_sequence(path, bars)
    assert [row["bar_index"] for row in sequence] == [2, 0, 1]
    assert [row["installation_step"] for row in sequence] == [1, 2, 3]
    assert stats["sequence_order_field"] == "row_order"



def test_manual_xlsx_name_column_accepts_bim_name_tail(tmp_path: Path):
    bars = [
        make_bar(0, [[0, 0, 0], [100, 0, 0]]),
        make_bar(1, [[0, 100, 0], [100, 100, 0]]),
    ]
    bars[0].name = "main:N4a:640520"
    bars[1].name = "main:N4a:640522"
    path = tmp_path / "name_tail_sequence.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name"])
    sheet.append(["640522"])
    sheet.append(["640520"])
    workbook.save(path)

    sequence, stats = load_manual_sequence(path, bars)
    assert [row["bar_index"] for row in sequence] == [1, 0]
    assert stats["sequence_order_field"] == "row_order"


def test_manual_xlsx_preinstalled_bars_are_step_zero(tmp_path: Path):
    bars = [
        make_bar(0, [[0, 0, 0], [100, 0, 0]]),
        make_bar(1, [[0, 100, 0], [100, 100, 0]]),
        make_bar(2, [[0, 200, 0], [100, 200, 0]]),
    ]
    path = tmp_path / "preinstalled_sequence.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "installation_status"])
    sheet.append(["bar-2", "待安装"])
    sheet.append(["bar-0", "已安装"])
    sheet.append(["bar-1", "待安装"])
    workbook.save(path)

    sequence, stats = load_manual_sequence(path, bars)
    assert [row["bar_index"] for row in sequence] == [0, 2, 1]
    assert [row["installation_step"] for row in sequence] == [0, 1, 2]
    assert [row["preinstalled"] for row in sequence] == [True, False, False]
    assert stats["preinstalled_bar_count"] == 1
    assert stats["pending_bar_count"] == 2


def test_preinstalled_bar_blocks_first_simulated_step():
    fixed = make_bar(0, [[0, 0, 0], [100, 0, 0]])
    moving = make_bar(1, [[0, 0, 0], [100, 0, 0]])
    world = CapsuleCollisionWorld([fixed, moving], {0: 0, 1: 1}, 1.0)
    contacts = world.contacts(moving, moving.axis, step=1)
    assert 0 in contacts

def test_automatic_topology_uses_multiple_candidate_axes():
    bars = [
        make_bar(0, [[0, 0, 0], [100, 0, 0]]),
        make_bar(1, [[0, 100, 100], [100, 100, 100]]),
    ]
    sequence, stats = plan_installation(bars, ["z", "y", "x"], 1.0)
    assert len(sequence) == 2
    assert stats["candidate_axes"] == ["Z", "Y", "X"]
    assert stats["tested_direction_count"] == 6
    assert all(abs(np.linalg.norm(row["entry_direction"]) - 1.0) < 1e-8 for row in sequence)


def test_se3_collision_planner_outputs_safe_paths(tmp_path: Path):
    bars = [
        make_bar(0, [[0, 0, 0], [100, 0, 0]]),
        make_bar(1, [[0, 100, 0], [100, 100, 0]]),
    ]
    sequence = [
        {
            "installation_step": index + 1,
            "bar_index": index,
            "entry_direction": [0.0, -1.0, 0.0],
        }
        for index in range(2)
    ]
    config = {
        "clearance_mm": 1.0,
        "grasp_fraction": 0.5,
        "outside_margin_mm": 100.0,
        "assembly_translation_step_mm": 25.0,
        "assembly_rotation_step_deg": 10.0,
        "assembly_rrt_iterations": 0,
        "assembly_random_seed": 3,
    }
    paths, summary = plan_assembly_paths(tmp_path, bars, sequence, config)
    assert summary["all_paths_collision_free"]
    assert summary["rigid_body_dof"] == 6
    assert set(paths) == {0, 1}
    assert (tmp_path / "assembly_paths.json").is_file()
    assert (tmp_path / "collision_report.json").is_file()
    with (tmp_path / "assembly_path_waypoints.csv").open(encoding="utf-8-sig") as stream:
        assert len(list(csv.DictReader(stream))) >= 4


def test_se3_planner_skips_preinstalled_bar_but_keeps_it_as_obstacle(tmp_path: Path):
    bars = [
        make_bar(0, [[0, 0, 0], [100, 0, 0]]),
        make_bar(1, [[0, 100, 0], [100, 100, 0]]),
    ]
    sequence = [
        {
            "installation_step": 0,
            "bar_index": 0,
            "entry_direction": [0.0, -1.0, 0.0],
            "preinstalled": True,
            "installation_status": "preinstalled",
        },
        {
            "installation_step": 1,
            "bar_index": 1,
            "entry_direction": [0.0, -1.0, 0.0],
            "preinstalled": False,
            "installation_status": "pending",
        },
    ]
    config = {
        "clearance_mm": 1.0,
        "grasp_fraction": 0.5,
        "outside_margin_mm": 100.0,
        "assembly_translation_step_mm": 25.0,
        "assembly_rotation_step_deg": 10.0,
        "assembly_rrt_iterations": 0,
        "assembly_random_seed": 3,
    }
    paths, summary = plan_assembly_paths(tmp_path, bars, sequence, config)
    assert set(paths) == {1}
    assert summary["preinstalled_bar_count"] == 1
    assert summary["simulated_bar_count"] == 1


def test_viewer_starts_with_preinstalled_bars_and_animates_pending_only(tmp_path: Path):
    bars = [
        make_bar(0, [[0, 0, 0], [100, 0, 0]]),
        make_bar(1, [[0, 100, 0], [100, 100, 0]]),
    ]
    sequence = []
    for step, bar, preinstalled in [(0, bars[0], True), (1, bars[1], False)]:
        sequence.append({
            "installation_step": step,
            "bar_index": bar.index,
            "entity_id": bar.entity_id,
            "guid": bar.guid,
            "name": bar.name,
            "tag": bar.tag,
            "length_mm": bar.length,
            "radius_mm": bar.radius,
            "entry_direction": [0.0, -1.0, 0.0],
            "forced_core_resolution": False,
            "preinstalled": preinstalled,
            "installation_status": "preinstalled" if preinstalled else "pending",
        })
    save_planning_outputs(tmp_path, bars, {}, sequence, {}, {})
    viewer = json.loads((tmp_path / "viewer_model.json").read_text(encoding="utf-8"))
    assert viewer["initial_installed"] == [0]
    assert [row["i"] for row in viewer["sequence"]] == [1]


def test_robot_export_excludes_preinstalled_bars(tmp_path: Path):
    bars = [
        make_bar(0, [[0, 0, 0], [100, 0, 0]]),
        make_bar(1, [[0, 100, 0], [100, 100, 0]]),
    ]
    sequence = [
        {"installation_step": 0, "bar_index": 0, "entry_direction": [0, -1, 0], "preinstalled": True},
        {"installation_step": 1, "bar_index": 1, "entry_direction": [0, -1, 0], "preinstalled": False},
    ]
    config = {
        "robot_sample_period_s": 0.1,
        "robot_linear_speed_mm_s": 1000.0,
        "robot_angular_speed_deg_s": 90.0,
        "outside_margin_mm": 100.0,
        "preinsert_distance_mm": 25.0,
        "retreat_distance_mm": 25.0,
        "grasp_fraction": 0.5,
    }
    summary = generate_robot_outputs(tmp_path, bars, sequence, config)
    preview = json.loads((tmp_path / "robot_waypoints.json").read_text(encoding="utf-8"))
    assert summary["preinstalled_bar_count"] == 1
    assert summary["requested_bar_count"] == 1
    assert [row["bar_index"] for row in preview] == [1]
