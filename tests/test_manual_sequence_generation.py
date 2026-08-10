from pathlib import Path

import numpy as np
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import main
from app.services.ifc_geometry import Rebar
from app.services.sequence_io import load_manual_sequence


def _bar(index: int, z: float, tag: str) -> Rebar:
    axis = np.asarray([[0.0, 0.0, z], [100.0, 0.0, z]])
    return Rebar(
        index=index,
        entity_id=100 + index,
        guid=f"guid-{index}",
        name=f"main_N{index}:main_N{index}:{tag}",
        tag=tag,
        map_id=0,
        axis=axis,
        radius=5.0,
        bbox_min=axis.min(0) - 5.0,
        bbox_max=axis.max(0) + 5.0,
        length=100.0,
    )


def test_generate_manual_sequence_endpoint(monkeypatch):
    bars = [_bar(0, -100.0, "640520"), _bar(1, 0.0, "640522"), _bar(2, 100.0, "715563")]
    monkeypatch.setattr(main, "parse_ifc_rebars", lambda path: (bars, {}, {"ifc_schema": "IFC2X3"}))
    client = TestClient(main.app)
    response = client.post(
        "/api/sequence/generate",
        files={"file": ("model.ifc", b"ISO-10303-21;", "application/octet-stream")},
    )
    assert response.status_code == 200
    assert response.content[:2] == b"PK"
    output = Path("generated_sequence_test.xlsx")
    output.write_bytes(response.content)
    try:
        workbook = load_workbook(output, data_only=False)
        sheet = workbook.worksheets[0]
        assert [sheet.cell(row, 1).value for row in range(2, 5)] == ["640520", "640522", "715563"]
        assert sheet.cell(1, 1).value == "name"
        assert sheet.cell(1, 2).value == "installation_status"
        assert [sheet.cell(row, 2).value for row in range(2, 5)] == ["待安装"] * 3
        sheet.cell(2, 2).value = "已安装"
        assert workbook.worksheets[1]["B4"].value.startswith("=COUNTA")
        workbook.active = 1
        workbook.save(output)
        sequence, stats = load_manual_sequence(output, bars)
        assert [item["tag"] for item in sequence] == ["640520", "640522", "715563"]
        assert [item["installation_step"] for item in sequence] == [0, 1, 2]
        assert [item["preinstalled"] for item in sequence] == [True, False, False]
        assert stats["preinstalled_bar_count"] == 1
        assert stats["pending_bar_count"] == 2
        assert stats["sequence_order_field"] == "row_order"
    finally:
        output.unlink(missing_ok=True)

