from fastapi.testclient import TestClient
from io import BytesIO
import json
from openpyxl import load_workbook
import app.main as main_module
from app.main import app


def test_health():
    client=TestClient(app)
    r=client.get('/api/health')
    assert r.status_code==200
    assert r.json()['status']=='ok'
    assert r.json()['version']=='1.9.0'


def test_index():
    client=TestClient(app)
    r=client.get('/')

    assert r.status_code==200
    assert '钢筋空间拓扑规划与碰撞检测系统' in r.text
    assert 'id="motionGuideToggle"' in r.text
    assert 'id="poseAxesToggle"' in r.text
    assert 'id="motionHud"' in r.text
    assert 'value="visual"' in r.text
    assert 'value="visual_groups"' in r.text
    assert 'id="visualSequenceEditor"' in r.text
    assert 'id="visualBoxSelectBtn"' in r.text
    assert 'id="meshGroupEditor"' in r.text
    assert 'id="meshInputMode"' in r.text
    assert 'id="meshConfirmGroupBtn"' in r.text
    assert 'id="meshSolveBtn"' in r.text
    assert 'id="meshPreviewSlider"' in r.text
    assert 'id="rerunMeshGroupBtn"' in r.text
    assert "框选多根时按模型索引依次追加" in r.text
    for speed in ("0.01", "0.03", "0.05", "0.1"):
        assert f'<option value="{speed}">{speed}×</option>' in r.text

def test_sequence_template():
    client=TestClient(app)
    r=client.get('/api/templates/installation-sequence')
    assert r.status_code==200
    assert r.content[:2]==b'PK'
    workbook = load_workbook(BytesIO(r.content), read_only=False)
    sheet = workbook["安装顺序"]
    assert [sheet.cell(1, col).value for col in range(1, 3)] == ["name", "installation_status"]
    assert len(sheet.data_validations.dataValidation) == 1
    workbook.close()


def test_visual_sequence_preview_parses_ifc():
    content = b"""ISO-10303-21;
HEADER;
FILE_SCHEMA(('IFC2X3'));
ENDSEC;
DATA;
#1=IFCCARTESIANPOINT((0.,0.,0.));
#2=IFCCARTESIANPOINT((100.,0.,0.));
#3=IFCPOLYLINE((#1,#2));
#4=IFCCOMPOSITECURVESEGMENT(.CONTINUOUS.,.T.,#3);
#5=IFCCOMPOSITECURVE((#4),.F.);
#6=IFCSWEPTDISKSOLID(#5,6.,$,0.,1.);
#7=IFCSHAPEREPRESENTATION($,'Body','AdvancedSweptSolid',(#6));
#8=IFCPRODUCTDEFINITIONSHAPE($,$,(#7));
#9=IFCREINFORCINGBAR('guid',$,'rebar:640520',$,$,$,#8,'640520',12.,113.,100.,.NOTDEFINED.,$);
ENDSEC;
END-ISO-10303-21;
"""
    client = TestClient(app)
    response = client.post(
        "/api/sequence/preview",
        files={"file": ("preview.ifc", content, "application/octet-stream")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["meta"]["rebar_count"] == 1
    assert payload["bars"][0]["n"] == "640520"
    assert payload["bars"][0]["p"] == [[0.0, 0.0, 0.0], [100.0, 0.0, 0.0]]


def _single_rebar_ifc(bim_id: str, y: float = 0.0) -> bytes:
    return f"""ISO-10303-21;
HEADER;
FILE_SCHEMA(('IFC2X3'));
ENDSEC;
DATA;
#1=IFCCARTESIANPOINT((0.,{y},0.));
#2=IFCCARTESIANPOINT((100.,{y},0.));
#3=IFCPOLYLINE((#1,#2));
#4=IFCCOMPOSITECURVESEGMENT(.CONTINUOUS.,.T.,#3);
#5=IFCCOMPOSITECURVE((#4),.F.);
#6=IFCSWEPTDISKSOLID(#5,6.,$,0.,1.);
#7=IFCSHAPEREPRESENTATION($,'Body','AdvancedSweptSolid',(#6));
#8=IFCPRODUCTDEFINITIONSHAPE($,$,(#7));
#9=IFCREINFORCINGBAR('guid-{bim_id}',$,'rebar:{bim_id}',$,$,$,#8,'{bim_id}',12.,113.,100.,.NOTDEFINED.,$);
ENDSEC;
END-ISO-10303-21;
""".encode()


def test_multiple_ifc_preview_creates_one_mesh_group_per_file():
    response = TestClient(app).post(
        "/api/sequence/preview",
        files=[
            ("file", ("顶板.ifc", _single_rebar_ifc("640520", 0), "application/octet-stream")),
            ("file", ("腹板.ifc", _single_rebar_ifc("640522", 100), "application/octet-stream")),
        ],
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["mesh_group_input_mode"] == "multiple_ifc_files"
    assert payload["source_filenames"] == ["顶板.ifc", "腹板.ifc"]
    assert payload["meta"]["source_file_count"] == 2
    assert [bar["i"] for bar in payload["bars"]] == [0, 1]
    assert [bar["n"] for bar in payload["bars"]] == ["640520", "640522"]
    assert payload["suggested_mesh_groups"] == [
        {
            "group_id": "G001", "name": "顶板", "installation_step": 1,
            "installation_status": "pending", "bar_indices": [0],
            "source_filename": "顶板.ifc",
        },
        {
            "group_id": "G002", "name": "腹板", "installation_step": 2,
            "installation_status": "pending", "bar_indices": [1],
            "source_filename": "腹板.ifc",
        },
    ]


def test_multiple_ifc_mesh_task_persists_all_source_models(monkeypatch, tmp_path):
    captured = {}
    monkeypatch.setattr(main_module, "JOBS_DIR", tmp_path)
    monkeypatch.setattr(
        main_module,
        "create_task",
        lambda task_id, filename, options: captured.update(
            task_id=task_id, filename=filename, options=options
        ),
    )
    monkeypatch.setattr(main_module, "launch_worker", lambda mode, task_id, config=None: None)
    group_payload = {
        "mode": "mesh_groups",
        "schema_version": 4,
        "motion_model": "pending_group_descent_then_cumulative_rotation",
        "input_mode": "multiple_ifc_files",
        "model_fingerprint": "sha256:combined-preview",
        "assembly_rotation_axis": None,
        "groups": [
            {"group_id": "G001", "installation_step": 1, "bar_indices": [0], "source_filename": "顶板.ifc"},
            {"group_id": "G002", "installation_step": 2, "bar_indices": [1], "source_filename": "腹板.ifc"},
        ],
    }
    response = TestClient(app).post(
        "/api/tasks",
        files=[
            ("file", ("顶板.ifc", b"TOP IFC", "application/octet-stream")),
            ("file", ("腹板.ifc", b"WEB IFC", "application/octet-stream")),
        ],
        data={
            "options_json": json.dumps({"sequence_source": "visual_groups"}),
            "visual_sequence_json": json.dumps(group_payload, ensure_ascii=False),
        },
    )

    assert response.status_code == 202
    job_dir = tmp_path / captured["task_id"]
    assert (job_dir / "input.ifc").read_bytes() == b"TOP IFC"
    assert (job_dir / "input_models" / "0002.ifc").read_bytes() == b"WEB IFC"
    manifest = json.loads((job_dir / "input_models.json").read_text(encoding="utf-8"))
    assert [item["source_filename"] for item in manifest["models"]] == ["顶板.ifc", "腹板.ifc"]
    assert captured["filename"] == "顶板.ifc 等 2 个网片 IFC"


def test_multiple_ifc_files_are_rejected_outside_mesh_group_mode():
    response = TestClient(app).post(
        "/api/tasks",
        files=[
            ("file", ("one.ifc", b"IFC 1", "application/octet-stream")),
            ("file", ("two.ifc", b"IFC 2", "application/octet-stream")),
        ],
        data={"options_json": json.dumps({"sequence_source": "automatic"})},
    )
    assert response.status_code == 400
    assert "多个 IFC 文件仅用于可视化网片组" in response.json()["detail"]


def test_visual_mesh_group_preview_resolves_shared_rigid_path():
    content = b"""ISO-10303-21;
HEADER;
FILE_SCHEMA(('IFC2X3'));
ENDSEC;
DATA;
#1=IFCCARTESIANPOINT((0.,0.,0.));
#2=IFCCARTESIANPOINT((100.,0.,0.));
#3=IFCPOLYLINE((#1,#2));
#4=IFCCOMPOSITECURVESEGMENT(.CONTINUOUS.,.T.,#3);
#5=IFCCOMPOSITECURVE((#4),.F.);
#6=IFCSWEPTDISKSOLID(#5,6.,$,0.,1.);
#7=IFCSHAPEREPRESENTATION($,'Body','AdvancedSweptSolid',(#6));
#8=IFCPRODUCTDEFINITIONSHAPE($,$,(#7));
#9=IFCREINFORCINGBAR('guid',$,'rebar:640520',$,$,$,#8,'640520',12.,113.,100.,.NOTDEFINED.,$);
ENDSEC;
END-ISO-10303-21;
"""
    client = TestClient(app)
    initial = client.post(
        "/api/sequence/preview",
        files={"file": ("preview.ifc", content, "application/octet-stream")},
    )
    assert initial.status_code == 200
    fingerprint = initial.json()["model_fingerprint"]
    groups = {
        "mode": "mesh_groups",
        "schema_version": 4,
        "motion_model": "pending_group_descent_then_cumulative_rotation",
        "model_fingerprint": fingerprint,
        "assembly_rotation_axis": {
            "transverse_mm": None,
            "elevation_mm": None,
            "direction": [1, 0, 0],
        },
        "staging_clearance_mm": 800,
        "groups": [{
            "group_id": "G001",
            "name": "顶板",
            "installation_step": 1,
            "installation_status": "pending",
            "bar_indices": [0],
            "plane_angle_deg": 0,
        }],
    }
    response = client.post(
        "/api/sequence/preview",
        files={"file": ("preview.ifc", content, "application/octet-stream")},
        data={"visual_sequence_json": json.dumps(groups, ensure_ascii=False)},
    )
    assert response.status_code == 200
    resolved = response.json()["mesh_groups"]
    assert resolved["mode"] == "mesh_groups"
    assert resolved["schema_version"] == 4
    assert resolved["motion_model"] == "pending_group_descent_then_cumulative_rotation"
    assert resolved["assembly_rotation_axis"]["direction"] == [1.0, 0.0, 0.0]
    assert resolved["group_count"] == 1
    group = resolved["groups"][0]
    assert group["bar_indices"] == [0]
    assert group["minimum_staging_clearance_mm"] == 800.0
    assert [phase["name"] for phase in resolved["paths"][0]["phases"]] == [
        "pending_group_descent",
        "installed_assembly_rotation_to_next",
    ]
    assert resolved["paths"][0]["phases"][0]["collision_checked"] is False
    assert resolved["initial_preparation"]["omitted"] is True
    assert resolved["final_restore"]["collision_checked"] is False


def test_visual_mesh_group_preview_rejects_wrong_model_fingerprint():
    content = b"""ISO-10303-21;
HEADER;
FILE_SCHEMA(('IFC2X3'));
ENDSEC;
DATA;
#1=IFCCARTESIANPOINT((0.,0.,0.));
#2=IFCCARTESIANPOINT((100.,0.,0.));
#3=IFCPOLYLINE((#1,#2));
#4=IFCCOMPOSITECURVESEGMENT(.CONTINUOUS.,.T.,#3);
#5=IFCCOMPOSITECURVE((#4),.F.);
#6=IFCSWEPTDISKSOLID(#5,6.,$,0.,1.);
#7=IFCSHAPEREPRESENTATION($,'Body','AdvancedSweptSolid',(#6));
#8=IFCPRODUCTDEFINITIONSHAPE($,$,(#7));
#9=IFCREINFORCINGBAR('guid',$,'rebar:640520',$,$,$,#8,'640520',12.,113.,100.,.NOTDEFINED.,$);
ENDSEC;
END-ISO-10303-21;
"""
    group_payload = {
        "mode": "mesh_groups", "schema_version": 4,
        "motion_model": "pending_group_descent_then_cumulative_rotation",
        "model_fingerprint": "sha256:not-this-model",
        "assembly_rotation_axis": None,
        "groups": [{"group_id": "G1", "installation_step": 1, "bar_indices": [0]}],
    }
    response = TestClient(app).post(
        "/api/sequence/preview",
        files={"file": ("preview.ifc", content, "application/octet-stream")},
        data={"visual_sequence_json": json.dumps(group_payload)},
    )
    assert response.status_code == 422
    assert "不同的 IFC" in response.json()["detail"]


def test_visual_mesh_group_task_forces_collision_and_disables_robot(monkeypatch, tmp_path):
    captured = {}
    monkeypatch.setattr(main_module, "JOBS_DIR", tmp_path)
    monkeypatch.setattr(
        main_module,
        "create_task",
        lambda task_id, filename, options: captured.update(
            task_id=task_id, filename=filename, options=options
        ),
    )
    monkeypatch.setattr(main_module, "launch_worker", lambda mode, task_id, config=None: None)
    group_payload = {
        "mode": "mesh_groups",
        "schema_version": 4,
        "motion_model": "pending_group_descent_then_cumulative_rotation",
        "model_fingerprint": "sha256:preview-fingerprint",
        "assembly_rotation_axis": None,
        "groups": [{"group_id": "G1", "installation_step": 1, "bar_indices": [0]}],
    }
    response = TestClient(app).post(
        "/api/tasks",
        files={"file": ("model.ifc", b"IFC", "application/octet-stream")},
        data={
            "options_json": json.dumps(
                {
                    "sequence_source": "visual_groups",
                    "generate_assembly_paths": False,
                    "generate_robot_path": True,
                }
            ),
            "visual_sequence_json": json.dumps(group_payload),
        },
    )
    assert response.status_code == 202
    assert captured["options"]["sequence_source"] == "visual_groups"
    assert captured["options"]["generate_assembly_paths"] is True
    assert captured["options"]["generate_robot_path"] is False
    saved = json.loads(
        (tmp_path / captured["task_id"] / "input_sequence.json").read_text(encoding="utf-8")
    )
    assert saved["mode"] == "mesh_groups"


def test_mesh_group_task_rejects_robot_regeneration(monkeypatch):
    monkeypatch.setattr(
        main_module,
        "get_task",
        lambda task_id: {
            "id": task_id,
            "status": "completed",
            "options": {"sequence_source": "visual_groups"},
        },
    )
    response = TestClient(app).post(
        "/api/tasks/group-task/robot",
        json={
            "linear_speed_mm_s": 600,
            "angular_speed_deg_s": 45,
            "sample_period_s": 0.1,
            "outside_margin_mm": 800,
            "preinsert_distance_mm": 250,
            "retreat_distance_mm": 300,
            "grasp_fraction": 0.5,
        },
    )
    assert response.status_code == 409
    assert "多点夹具" in response.json()["detail"]


def test_completed_mesh_group_task_can_rerun_without_regrouping(monkeypatch, tmp_path):
    source_id = "completed-group-task"
    source_dir = tmp_path / source_id
    source_dir.mkdir()
    (source_dir / "input.ifc").write_bytes(b"saved IFC")
    group_payload = {
        "mode": "mesh_groups",
        "schema_version": 2,
        "model_fingerprint": "sha256:saved-model",
        "top_elevation_mm": 5000,
        "rotation_axis": {"transverse_mm": 100, "elevation_mm": 200},
        "groups": [{
            "group_id": "G1",
            "name": "腹板",
            "installation_step": 1,
            "installation_status": "preinstalled",
            "bar_indices": [0],
            "plane_angle_deg": 90,
            "rotation_axis": {
                "transverse_mm": 123,
                "elevation_mm": 456,
                "direction": [1, 0, 0],
            },
            "staging_clearance_mm": 950,
        }],
    }
    (source_dir / "input_sequence.json").write_text(
        json.dumps(group_payload), encoding="utf-8"
    )
    source_task = {
        "id": source_id,
        "filename": "saved.ifc",
        "status": "completed",
        "options": {
            "sequence_source": "visual_groups",
            "clearance_mm": 2.0,
            "assembly_translation_step_mm": 25.0,
            "assembly_rotation_step_deg": 2.5,
        },
    }
    created = {}
    launched = {}
    monkeypatch.setattr(main_module, "JOBS_DIR", tmp_path)
    monkeypatch.setattr(
        main_module, "get_task", lambda task_id: source_task if task_id == source_id else None
    )
    monkeypatch.setattr(
        main_module,
        "create_task",
        lambda task_id, filename, options: created.update(
            task_id=task_id, filename=filename, options=options
        ),
    )
    monkeypatch.setattr(
        main_module,
        "launch_worker",
        lambda mode, task_id, config=None: launched.update(mode=mode, task_id=task_id),
    )

    response = TestClient(app).post(f"/api/tasks/{source_id}/rerun")

    assert response.status_code == 202
    result = response.json()
    assert result["source_task_id"] == source_id
    assert result["reused_mesh_groups"] is True
    assert result["schema_version"] == 4
    assert result["migrated_from_schema_version"] == 2
    assert result["task_id"] != source_id
    assert created["filename"] == "saved.ifc"
    assert created["options"]["sequence_source"] == "visual_groups"
    assert created["options"]["generate_assembly_paths"] is True
    assert created["options"]["generate_robot_path"] is False
    assert launched == {"mode": "planning", "task_id": result["task_id"]}
    new_dir = tmp_path / result["task_id"]
    assert (new_dir / "input.ifc").read_bytes() == b"saved IFC"
    migrated = json.loads((new_dir / "input_sequence.json").read_text(encoding="utf-8"))
    assert migrated["schema_version"] == 4
    assert migrated["motion_model"] == "pending_group_descent_then_cumulative_rotation"
    assert migrated["longitudinal_axis"] == "auto"
    assert migrated["assembly_rotation_axis"] == {
        "transverse_mm": None,
        "elevation_mm": None,
        "direction": None,
    }
    assert "top_elevation_mm" not in migrated
    assert "rotation_axis" not in migrated["groups"][0]
    assert migrated["groups"][0] == {
        "group_id": "G1",
        "name": "腹板",
        "installation_step": 1,
        "installation_status": "preinstalled",
        "bar_indices": [0],
        "plane_angle_deg": 90,
        "staging_clearance_mm": 950,
    }
    assert (source_dir / "input_sequence.json").is_file()
    assert json.loads(source_dir.joinpath("input_sequence.json").read_text(encoding="utf-8")) == group_payload


def test_v3_migration_preserves_manual_longitudinal_and_shared_rotation_axis():
    manual_axis = {
        "point_mm": [10.0, -25.0, 3600.0],
        "transverse_mm": -25.0,
        "elevation_mm": 3600.0,
        "direction": [0.70710678, 0.70710678, 0.0],
        "source": "manual",
    }
    legacy = {
        "mode": "mesh_groups",
        "schema_version": 3,
        "motion_model": "cumulative_installed_rotation_then_pending_descent",
        "model_fingerprint": "sha256:v3-model",
        "longitudinal_axis": [0.70710678, 0.70710678, 0.0],
        "vertical_axis": [0, 0, 1],
        "staging_clearance_mm": 825,
        "assembly_rotation_axis": manual_axis,
        "groups": [{
            "group_id": "G1",
            "name": "倒角",
            "installation_step": 1,
            "installation_status": "pending",
            "bar_indices": [0],
            "plane_angle_deg": -45,
            "minimum_staging_clearance_mm": 900,
        }],
    }

    migrated, migrated_from = main_module._migrate_mesh_group_payload(legacy)

    assert migrated_from == 3
    assert migrated["schema_version"] == 4
    assert migrated["motion_model"] == "pending_group_descent_then_cumulative_rotation"
    assert migrated["longitudinal_axis"] == legacy["longitudinal_axis"]
    assert migrated["assembly_rotation_axis"] == manual_axis
    assert migrated["groups"][0]["plane_angle_deg"] == -45
    assert migrated["groups"][0]["staging_clearance_mm"] == 900


def test_new_mesh_group_task_rejects_legacy_v2_payload(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module, "JOBS_DIR", tmp_path)
    payload = {
        "mode": "mesh_groups",
        "schema_version": 2,
        "model_fingerprint": "sha256:legacy",
        "groups": [{"group_id": "G1", "installation_step": 1, "bar_indices": [0]}],
    }
    response = TestClient(app).post(
        "/api/tasks",
        files={"file": ("model.ifc", b"IFC", "application/octet-stream")},
        data={
            "options_json": json.dumps({"sequence_source": "visual_groups"}),
            "visual_sequence_json": json.dumps(payload),
        },
    )
    assert response.status_code == 422
    assert "schema_version 必须是 4" in response.json()["detail"]


def test_non_mesh_group_task_cannot_use_mesh_group_rerun(monkeypatch):
    monkeypatch.setattr(
        main_module,
        "get_task",
        lambda task_id: {
            "id": task_id,
            "status": "completed",
            "options": {"sequence_source": "automatic"},
        },
    )

    response = TestClient(app).post("/api/tasks/automatic-task/rerun")

    assert response.status_code == 409
    assert "可视化网片组" in response.json()["detail"]
