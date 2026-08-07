from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .ifc_geometry import Rebar


GROUP_BOTTOM = "底板"
GROUP_WEB = "腹板"
GROUP_TOP = "顶板"
GROUP_ORDER = {GROUP_BOTTOM: 0, GROUP_WEB: 1, GROUP_TOP: 2}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_SECONDARY_FILL = PatternFill("solid", fgColor="5B6770")
_BODY_FILL = PatternFill("solid", fgColor="F7FAFC")
_BOTTOM_FILL = PatternFill("solid", fgColor="E8F1FB")
_WEB_FILL = PatternFill("solid", fgColor="FFF4E5")
_TOP_FILL = PatternFill("solid", fgColor="EAF6EA")
_LABEL_FILL = PatternFill("solid", fgColor="EAF3F8")
_WHITE_FONT = Font(bold=True, color="FFFFFF")
_BOLD_FONT = Font(bold=True)
_THIN = Side(style="thin", color="D9E2EC")
_MEDIUM = Side(style="medium", color="6B8EAD")


def _family(name: str) -> str:
    first = (name or "").split(":", 1)[0]
    return first.rsplit("_", 1)[-1] if "_" in first else first


def _name_identifier(bar: Rebar) -> str:
    if bar.tag:
        return str(bar.tag).strip()
    name = str(bar.name or "").strip()
    return name.rsplit(":", 1)[-1].strip() or name


def _classify_groups(rebars: Iterable[Rebar]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    bars = list(rebars)
    if not bars:
        raise ValueError("IFC 中未识别到钢筋")
    centers = [float((bar.bbox_min[2] + bar.bbox_max[2]) * 0.5) for bar in bars]
    z_min = min(centers)
    z_max = max(centers)
    z_span = max(z_max - z_min, 0.0)
    # The generated grouping is only a starting suggestion.  It scales with
    # the uploaded model and keeps bars spanning most of the depth in the web.
    bottom_limit = z_min + 0.20 * z_span
    top_limit = z_max - 0.20 * z_span
    depth_limit = 0.50 * z_span

    records: list[dict[str, Any]] = []
    for bar in bars:
        center = float((bar.bbox_min[2] + bar.bbox_max[2]) * 0.5)
        bar_z_span = float(bar.bbox_max[2] - bar.bbox_min[2])
        if z_span > 1e-9 and center <= bottom_limit and bar_z_span <= depth_limit:
            group = GROUP_BOTTOM
            basis = f"z 中心 ≤ {bottom_limit:.1f} mm：底板建议"
        elif z_span > 1e-9 and center >= top_limit and bar_z_span <= depth_limit:
            group = GROUP_TOP
            basis = f"z 中心 ≥ {top_limit:.1f} mm：顶板建议"
        else:
            group = GROUP_WEB
            basis = "其余标高范围或跨越大部分高度：腹板建议"
        records.append(
            {
                "name": _name_identifier(bar),
                "construction_group": group,
                "model_name": bar.name,
                "model_bar_index": bar.index,
                "ifc_entity_id": bar.entity_id,
                "bim_tag": bar.tag,
                "z_center_mm": round(center, 3),
                "z_span_mm": round(bar_z_span, 3),
                "bar_family": _family(bar.name),
                "group_basis": basis,
            }
        )

    name_counts = Counter(row["name"] for row in records)
    for row in records:
        if name_counts[row["name"]] > 1:
            row["name"] = row["model_name"]
    if len({row["name"] for row in records}) != len(records):
        raise ValueError("IFC ??????? name ?????????????? BIM name ? guid")

    records.sort(
        key=lambda row: (
            GROUP_ORDER[row["construction_group"]],
            row["z_center_mm"],
            row["bar_family"],
            row["model_bar_index"],
        )
    )
    for order, row in enumerate(records, 1):
        row["display_step"] = order
    return records, {
        "rebar_count": len(records),
        "group_counts": dict(Counter(row["construction_group"] for row in records)),
        "z_center_min_mm": z_min,
        "z_center_max_mm": z_max,
        "bottom_limit_mm": bottom_limit,
        "top_limit_mm": top_limit,
        "group_rule": "按模型 z 中心标高分组：底部 20% 为底板，顶部 20% 为顶板，其余为腹板；跨越大部分高度的钢筋归入腹板。",
    }


def _style_header(row, fill: PatternFill = _HEADER_FILL) -> None:
    for cell in row:
        cell.fill = fill
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)


def _style_body(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.fill = _BODY_FILL
            cell.border = Border(bottom=_THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def _set_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def build_manual_sequence_workbook(
    rebars: Iterable[Rebar],
    output_path: Path,
    *,
    source_filename: str = "uploaded.ifc",
    ifc_meta: dict[str, Any] | None = None,
) -> dict[str, Any]:
    records, grouping = _classify_groups(rebars)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sequence = workbook.active
    sequence.title = "安装顺序"
    sequence.sheet_view.showGridLines = False
    sequence.freeze_panes = "A2"

    headers = [
        "name",
        "construction_group",
        "model_name",
        "model_bar_index",
        "ifc_entity_id",
        "bim_tag",
        "z_center_mm",
        "z_span_mm",
        "bar_family",
        "group_basis",
    ]
    sequence.append(headers)
    for record in records:
        sequence.append([record.get(header) for header in headers])
    _style_header(sequence[1])
    if records:
        _style_body(sequence, 2, len(records) + 1, 1, len(headers))
    sequence.auto_filter.ref = f"A1:J{len(records) + 1}"
    sequence.sheet_properties.pageSetUpPr.fitToPage = True
    sequence.page_setup.fitToWidth = 1
    sequence.page_setup.fitToHeight = 0
    sequence.column_dimensions["A"].width = 16
    sequence.column_dimensions["B"].width = 16
    sequence.column_dimensions["C"].width = 46
    sequence.column_dimensions["D"].width = 16
    sequence.column_dimensions["E"].width = 16
    sequence.column_dimensions["F"].width = 14
    sequence.column_dimensions["G"].width = 14
    sequence.column_dimensions["H"].width = 12
    sequence.column_dimensions["I"].width = 12
    sequence.column_dimensions["J"].width = 34
    sequence.row_dimensions[1].height = 30
    for row in range(2, len(records) + 2):
        sequence.cell(row, 1).number_format = "@"
        sequence.cell(row, 6).number_format = "@"
        sequence.cell(row, 7).number_format = "0.0"
        sequence.cell(row, 8).number_format = "0.0"
        sequence.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        sequence.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        for col in (4, 5, 7, 8):
            sequence.cell(row, col).alignment = Alignment(horizontal="right", vertical="center")
    group_ends: dict[str, int] = {}
    for row, record in enumerate(records, 2):
        group_ends[record["construction_group"]] = row
    fills = {GROUP_BOTTOM: _BOTTOM_FILL, GROUP_WEB: _WEB_FILL, GROUP_TOP: _TOP_FILL}
    previous_group = None
    for row, record in enumerate(records, 2):
        group = record["construction_group"]
        sequence.cell(row, 2).fill = fills[group]
        if previous_group is not None and group != previous_group:
            for cell in sequence[row - 1]:
                cell.border = Border(bottom=_MEDIUM)
        previous_group = group
    if records:
        for cell in sequence[len(records) + 1]:
            cell.border = Border(bottom=_MEDIUM)

    check = workbook.create_sheet("校核")
    check.sheet_view.showGridLines = False
    check.merge_cells("A1:D1")
    check["A1"] = "IFC 人工安装顺序校核"
    _style_header(check[1])
    check["A3"] = "校核项目"
    check["B3"] = "结果"
    _style_header(check[3], _SECONDARY_FILL)
    last_row = len(records) + 1
    check_rows = [
        ("钢筋总数", f"=COUNTA('安装顺序'!$A$2:$A${last_row})"),
        (f"{GROUP_BOTTOM}数量", f'=COUNTIF(\'安装顺序\'!$B$2:$B${last_row},"{GROUP_BOTTOM}")'),
        (f"{GROUP_WEB}数量", f'=COUNTIF(\'安装顺序\'!$B$2:$B${last_row},"{GROUP_WEB}")'),
        (f"{GROUP_TOP}数量", f'=COUNTIF(\'安装顺序\'!$B$2:$B${last_row},"{GROUP_TOP}")'),
        ("顺序完整性", '=IF(AND(B4>0,B5>0,B6>0),"顺序完整","需检查")'),
        ("安装顺序", f"{GROUP_BOTTOM} → {GROUP_WEB} → {GROUP_TOP}"),
    ]
    for row, (label, value) in enumerate(check_rows, 4):
        check.cell(row, 1, label)
        if isinstance(value, str) and value.startswith("="):
            check.cell(row, 2).value = value
        else:
            check.cell(row, 2, value)
        check.cell(row, 1).fill = _LABEL_FILL
        check.cell(row, 1).font = _BOLD_FONT
        for cell in check[row][:2]:
            cell.border = Border(bottom=_THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    check["A11"] = "来源文件"
    check["B11"] = source_filename
    check["A12"] = "IFC schema"
    check["B12"] = (ifc_meta or {}).get("ifc_schema", "unknown")
    check["A13"] = "实体数"
    check["B13"] = (ifc_meta or {}).get("entity_count", "")
    check["A14"] = "表示映射数"
    check["B14"] = (ifc_meta or {}).get("representation_map_count", "")
    check["A15"] = "分组依据"
    check["B15"] = grouping["group_rule"]
    for row in range(11, 16):
        check.cell(row, 1).fill = _LABEL_FILL
        check.cell(row, 1).font = _BOLD_FONT
        for cell in check[row][:2]:
            cell.border = Border(bottom=_THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    _set_widths(check, [24, 96, 18, 18])
    check.freeze_panes = "A3"
    check.row_dimensions[1].height = 28

    instructions = workbook.create_sheet("填写说明")
    instructions.sheet_view.showGridLines = False
    instructions.merge_cells("A1:B1")
    instructions["A1"] = "人工钢筋安装顺序填写说明"
    _style_header(instructions[1])
    instructions_rows = [
        ("字段", "说明"),
        ("name（必填）", "第一张“安装顺序”表中的 name 列是系统实际读取的字段。每行填一个 IFC 对应的 BIM 编号，例如 640520。"),
        ("行顺序", "第 2 行表示第 1 根安装，第 3 行表示第 2 根安装；系统按 name 列的行顺序读取。"),
        ("本表其他列", "其余列只用于核对模型信息，系统上传时只依赖 name 列；如调整顺序，请整行一起移动。"),
        ("覆盖要求", f"必须保留全部 {len(records)} 根钢筋，每个 BIM 编号只出现一次；系统会自动检查缺失和重复。"),
        ("默认分组", f"本表已按 {GROUP_BOTTOM} → {GROUP_WEB} → {GROUP_TOP} 生成，分组只是建议，可通过调整 name 行顺序自由修改。"),
        ("上传方式", "在系统中选择“Excel 人工顺序”，上传修改后的文件；之后系统会继续执行六自由度安装路径碰撞检查。"),
    ]
    for row_index, (label, value) in enumerate(instructions_rows, 3):
        instructions.cell(row_index, 1, label)
        instructions.cell(row_index, 2, value)
    _style_header(instructions[3], _SECONDARY_FILL)
    _style_body(instructions, 4, 3 + len(instructions_rows) - 1, 1, 2)
    for row in range(4, 3 + len(instructions_rows)):
        instructions.cell(row, 1).fill = _LABEL_FILL
        instructions.cell(row, 1).font = _BOLD_FONT
        instructions.row_dimensions[row].height = 38
    _set_widths(instructions, [22, 104])
    instructions.freeze_panes = "A3"
    instructions.row_dimensions[1].height = 28

    workbook.properties.title = "IFC 钢筋人工安装顺序"
    workbook.properties.subject = "可编辑的底板、腹板、顶板钢筋安装顺序"
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output_path)
    return {**grouping, "source_filename": source_filename, "output_path": str(output_path)}

