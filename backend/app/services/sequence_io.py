from __future__ import annotations

import csv
import math
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from openpyxl import load_workbook

from .ifc_geometry import Rebar


HEADER_ALIASES = {
    "installation_step": {"installation_step", "step", "order", "\u5b89\u88c5\u987a\u5e8f", "\u5b89\u88c5\u5e8f\u53f7", "\u5e8f\u53f7"},
    "installation_status": {
        "installation_status", "install_status", "status", "preinstalled", "installed",
        "\u5b89\u88c5\u72b6\u6001", "\u662f\u5426\u5df2\u5b89\u88c5", "\u521d\u59cb\u72b6\u6001",
    },
    "bar_id": {"bar_id", "rebar_id", "id", "\u94a2\u7b4bid"},
    "bar_index": {"bar_index", "index", "\u94a2\u7b4b\u7d22\u5f15", "\u6a21\u578b\u7d22\u5f15"},
    "entity_id": {"entity_id", "ifc_entity_id", "ifc\u5b9e\u4f53id", "\u5b9e\u4f53id"},
    "guid": {"guid", "ifc_guid", "ifcguid"},
    "tag": {"tag", "\u94a2\u7b4b\u7f16\u53f7", "\u7f16\u53f7"},
    "name": {"name", "\u94a2\u7b4b\u540d\u79f0", "\u540d\u79f0"},
    "direction_x": {"direction_x", "entry_x", "\u8fdb\u573a\u65b9\u5411x", "\u65b9\u5411x"},
    "direction_y": {"direction_y", "entry_y", "\u8fdb\u573a\u65b9\u5411y", "\u65b9\u5411y"},
    "direction_z": {"direction_z", "entry_z", "\u8fdb\u573a\u65b9\u5411z", "\u65b9\u5411z"},
}

IDENTIFIER_COLUMNS = ("bar_index", "entity_id", "guid", "tag", "name")
GENERIC_IDENTIFIER_COLUMNS = ("bar_id", *IDENTIFIER_COLUMNS)

PREINSTALLED_VALUES = {
    "1", "true", "yes", "y", "installed", "preinstalled", "complete", "completed",
    "\u662f", "\u5df2\u5b89\u88c5", "\u5df2\u5b8c\u6210", "\u5b8c\u6210",
}
PENDING_VALUES = {
    "", "0", "false", "no", "n", "pending", "not_installed", "uninstalled",
    "\u5426", "\u5f85\u5b89\u88c5", "\u672a\u5b89\u88c5", "\u672a\u5b8c\u6210",
}


def _normal_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("-", "_")


def _canonical_headers(values: Iterable[Any]) -> dict[int, str]:
    aliases = {alias: key for key, names in HEADER_ALIASES.items() for alias in names}
    return {
        index: aliases[token]
        for index, value in enumerate(values)
        if (token := _normal_header(value)) in aliases
    }


def _read_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        active_headers = _canonical_headers(next(sheet.iter_rows(values_only=True), ()))
        if not active_headers:
            for candidate in workbook.worksheets:
                candidate_headers = _canonical_headers(next(candidate.iter_rows(values_only=True), ()))
                if candidate_headers:
                    sheet = candidate
                    break
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = _canonical_headers(next(iterator))
        except StopIteration:
            return []
        rows = [
            {name: values[index] if index < len(values) else None for index, name in headers.items()}
            for values in iterator
        ]
        workbook.close()
        return rows
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.reader(stream, delimiter=delimiter)
        try:
            headers = _canonical_headers(next(reader))
        except StopIteration:
            return []
        return [
            {name: values[index] if index < len(values) else None for index, name in headers.items()}
            for values in reader
        ]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_preinstalled(value: Any, sheet_row: int) -> bool:
    token = _clean_text(value).lower().replace(" ", "").replace("-", "_")
    if token in PREINSTALLED_VALUES:
        return True
    if token in PENDING_VALUES:
        return False
    raise ValueError(
        f"\u7b2c {sheet_row} \u884c installation_status \u5fc5\u987b\u662f\u2018\u5f85\u5b89\u88c5\u2019\u6216\u2018\u5df2\u5b89\u88c5\u2019"
    )


def _manual_default_direction(bar: Rebar, model_center: np.ndarray) -> list[float]:
    # Approach from the nearest outside half-space toward the model center.
    radial_out = bar.axis.mean(0) - model_center
    if np.linalg.norm(radial_out) < 1e-9:
        spans = bar.bbox_max - bar.bbox_min
        radial_out = np.zeros(3)
        radial_out[int(np.argmin(spans))] = 1.0
    entry = -radial_out / max(float(np.linalg.norm(radial_out)), 1e-12)
    return entry.tolist()


def _lookup_tokens(key: str, value: Any) -> list[str]:
    token = _clean_text(value)
    if not token:
        return []
    tokens = [token]
    if key == "name":
        tail = token.rsplit(":", 1)[-1].strip()
        if tail and tail != token:
            tokens.append(tail)
    return tokens


def _build_lookups(rebars: list[Rebar]) -> dict[str, dict[str, list[Rebar]]]:
    lookups: dict[str, dict[str, list[Rebar]]] = {key: {} for key in IDENTIFIER_COLUMNS}
    for bar in rebars:
        values = {
            "bar_index": str(bar.index),
            "entity_id": str(bar.entity_id),
            "guid": bar.guid,
            "tag": bar.tag,
            "name": bar.name,
        }
        for key, value in values.items():
            for token in _lookup_tokens(key, value):
                lookups[key].setdefault(token, []).append(bar)
    return lookups


def _match_row_to_bars(
    row: dict[str, Any],
    lookups: dict[str, dict[str, list[Rebar]]],
) -> set[int] | None:
    matches: set[int] | None = None
    for key in IDENTIFIER_COLUMNS:
        token = _clean_text(row.get(key))
        if not token:
            continue
        candidates = {bar.index for bar in lookups[key].get(token, [])}
        matches = candidates if matches is None else matches & candidates

    generic_token = _clean_text(row.get("bar_id"))
    if generic_token:
        generic_matches: set[int] = set()
        for key in IDENTIFIER_COLUMNS:
            generic_matches.update(bar.index for bar in lookups[key].get(generic_token, []))
        matches = generic_matches if matches is None else matches & generic_matches

    return matches


def load_manual_sequence(path: Path, rebars: list[Rebar]) -> tuple[list[dict], dict]:
    rows = _read_rows(path)
    if not rows:
        raise ValueError("\u4eba\u5de5\u5b89\u88c5\u987a\u5e8f\u8868\u4e3a\u7a7a\uff0c\u6216\u9996\u884c\u6ca1\u6709\u53ef\u8bc6\u522b\u7684\u5217\u540d")
    has_explicit_step = "installation_step" in rows[0]
    has_installation_status = "installation_status" in rows[0]
    if not set(GENERIC_IDENTIFIER_COLUMNS).intersection(rows[0]):
        raise ValueError("\u4eba\u5de5\u5b89\u88c5\u987a\u5e8f\u8868\u81f3\u5c11\u9700\u8981 bar_id\uff08\u94a2\u7b4bID\uff09\uff0c\u6216 bar_index\u3001entity_id\u3001guid\u3001tag\u3001name \u4e2d\u7684\u4e00\u5217")

    lookups = _build_lookups(rebars)
    by_index = {bar.index: bar for bar in rebars}

    parsed: list[tuple[int, Rebar, list[float] | None, int, bool]] = []
    for sheet_row, row in enumerate(rows, 2):
        if not any(_clean_text(value) for value in row.values()):
            continue
        preinstalled = _parse_preinstalled(row.get("installation_status"), sheet_row)
        if has_explicit_step and not preinstalled:
            try:
                step = int(float(_clean_text(row.get("installation_step"))))
            except Exception as exc:
                raise ValueError(f"\u7b2c {sheet_row} \u884c\u5b89\u88c5\u987a\u5e8f\u4e0d\u662f\u6574\u6570") from exc
            if step <= 0:
                raise ValueError(f"\u7b2c {sheet_row} \u884c\u5b89\u88c5\u987a\u5e8f\u5fc5\u987b\u5927\u4e8e 0")
        else:
            # The final pending sequence is renumbered after preinstalled rows
            # are separated, so sheet-row order is a stable temporary key.
            step = sheet_row - 1

        matches = _match_row_to_bars(row, lookups)
        if not matches:
            raise ValueError(f"\u7b2c {sheet_row} \u884c\u65e0\u6cd5\u5728 IFC \u4e2d\u552f\u4e00\u5b9a\u4f4d\u94a2\u7b4b")
        if len(matches) != 1:
            raise ValueError(f"\u7b2c {sheet_row} \u884c\u5339\u914d\u5230 {len(matches)} \u6839\u94a2\u7b4b\uff0c\u8bf7\u6539\u7528 guid \u6216 bar_index \u6d88\u9664\u6b67\u4e49")
        bar = by_index[next(iter(matches))]
        direction: list[float] | None = None
        raw_d = [row.get("direction_x"), row.get("direction_y"), row.get("direction_z")]
        if any(_clean_text(value) for value in raw_d):
            try:
                d = np.asarray([float(value) for value in raw_d], dtype=float)
            except Exception as exc:
                raise ValueError(f"\u7b2c {sheet_row} \u884c\u8fdb\u573a\u65b9\u5411\u5fc5\u987b\u540c\u65f6\u586b\u5199 X/Y/Z \u4e09\u4e2a\u6570\u5b57") from exc
            norm = float(np.linalg.norm(d))
            if not math.isfinite(norm) or norm < 1e-9:
                raise ValueError(f"\u7b2c {sheet_row} \u884c\u8fdb\u573a\u65b9\u5411\u4e0d\u80fd\u4e3a\u96f6\u5411\u91cf")
            direction = (d / norm).tolist()
        parsed.append((step, bar, direction, sheet_row, preinstalled))

    pending_parsed = [item for item in parsed if not item[4]]
    if len({item[0] for item in pending_parsed}) != len(pending_parsed):
        raise ValueError("\u4eba\u5de5\u5b89\u88c5\u987a\u5e8f\u5b58\u5728\u91cd\u590d\u7684 installation_step")
    indexes = [item[1].index for item in parsed]
    if len(set(indexes)) != len(indexes):
        raise ValueError("\u4eba\u5de5\u5b89\u88c5\u987a\u5e8f\u4e2d\u540c\u4e00\u6839\u94a2\u7b4b\u51fa\u73b0\u4e86\u591a\u6b21")
    missing = sorted(set(by_index) - set(indexes))
    if missing:
        example = ", ".join(str(value) for value in missing[:10])
        raise ValueError(f"\u4eba\u5de5\u5b89\u88c5\u987a\u5e8f\u7f3a\u5c11 {len(missing)} \u6839\u94a2\u7b4b\uff08\u4f8b\u5982 bar_index: {example}\uff09")

    preinstalled_parsed = sorted((item for item in parsed if item[4]), key=lambda item: item[3])
    pending_parsed.sort(key=lambda item: item[0])
    model_min = np.min([bar.bbox_min for bar in rebars], axis=0)
    model_max = np.max([bar.bbox_max for bar in rebars], axis=0)
    model_center = (model_min + model_max) * 0.5
    sequence: list[dict] = []
    ordered = [(0, item) for item in preinstalled_parsed]
    ordered.extend((new_step, item) for new_step, item in enumerate(pending_parsed, 1))
    for installation_step, (_, bar, direction, _, preinstalled) in ordered:
        sequence.append({
            "installation_step": installation_step,
            "bar_index": bar.index,
            "entity_id": bar.entity_id,
            "guid": bar.guid,
            "name": bar.name,
            "tag": bar.tag,
            "length_mm": bar.length,
            "radius_mm": bar.radius,
            "entry_direction": direction or _manual_default_direction(bar, model_center),
            "forced_core_resolution": False,
            "preinstalled": preinstalled,
            "installation_status": "preinstalled" if preinstalled else "pending",
        })
    return sequence, {
        "planner_mode": "manual_excel_sequence",
        "sequence_source": "excel",
        "row_count": len(sequence),
        "preinstalled_bar_count": len(preinstalled_parsed),
        "pending_bar_count": len(pending_parsed),
        "has_installation_status_column": has_installation_status,
        "sequence_order_field": "installation_step" if has_explicit_step else "row_order",
        "strict_graph_feasible": None,
        "certification_scope": (
            "user-supplied order; preinstalled bars are fixed obstacles and pending-bar "
            "feasibility is established only by downstream SE(3) collision checking"
        ),
    }
